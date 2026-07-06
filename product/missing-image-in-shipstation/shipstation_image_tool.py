#!/usr/bin/env python3
"""
ShipStation Product Image Tool v2
==================================
A local Flask web app for assigning images to ShipStation products.

Flow:
  1. Login with ShipStation API key + secret
  2. Fetches all products, filters to those missing an image
  3. Walks through them one-by-one with drag-and-drop from your website
  4. Pushes the image URL to ShipStation and moves to the next

Usage:
    pip install flask requests openpyxl beautifulsoup4
    python shipstation_image_tool.py

Then open http://localhost:5050 in your browser.
"""

import argparse
import base64
import io
import json
import os
import re
import time
import webbrowser
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from threading import Lock, Timer

from flask import Flask, render_template_string, request, jsonify, send_file
import requests
from requests.adapters import HTTPAdapter

# Reused session (connection pooling/keep-alive) for scraping cheshirehorse.com.
# Sized to comfortably cover several products' worth of concurrent PDP fetches.
scrape_session = requests.Session()
_scrape_adapter = HTTPAdapter(pool_connections=30, pool_maxsize=30)
scrape_session.mount("https://", _scrape_adapter)
scrape_session.mount("http://", _scrape_adapter)

# ─── Persistence ──────────────────────────────────────────────────────────────

PROGRESS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "progress.json")


def load_progress():
    """Load saved progress from disk. Returns dict with 'log' list."""
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"log": []}


log_lock = Lock()


def save_progress():
    """Write current log to disk."""
    data = {"log": state["log"], "saved_at": datetime.now().isoformat()}
    try:
        with open(PROGRESS_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except IOError as e:
        print(f"Warning: could not save progress: {e}")


def log_action(product, action, image_url="", source=None):
    """Append an action to the log and persist it. Locked because auto-match
    (background) and manual review can both be logging actions concurrently."""
    entry = {
        "sku": product.get("sku", ""),
        "name": product.get("name", ""),
        "url": image_url,
        "action": action,
        "timestamp": datetime.now().isoformat(),
    }
    if source:
        entry["source"] = source
    with log_lock:
        state["log"].append(entry)
        save_progress()


# ─── ShipStation API ──────────────────────────────────────────────────────────

API_BASE = "https://ssapi.shipstation.com"
RATE_LIMIT_DELAY = 1.6

api_lock = Lock()
last_request_time = 0


def ss_headers(api_key, api_secret):
    creds = base64.b64encode(f"{api_key}:{api_secret}".encode()).decode()
    return {"Authorization": f"Basic {creds}", "Content-Type": "application/json"}


def ss_request(method, path, api_key, api_secret, json_data=None, params=None):
    global last_request_time
    with api_lock:
        elapsed = time.time() - last_request_time
        if elapsed < RATE_LIMIT_DELAY:
            time.sleep(RATE_LIMIT_DELAY - elapsed)
        last_request_time = time.time()
    url = f"{API_BASE}{path}"
    headers = ss_headers(api_key, api_secret)
    resp = requests.request(method, url, headers=headers, json=json_data, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json() if resp.text else {}


def fetch_all_products(api_key, api_secret):
    products = []
    page = 1
    while True:
        data = ss_request("GET", "/products", api_key, api_secret,
                          params={"page": page, "pageSize": 500})
        batch = data.get("products", [])
        if not batch:
            break
        products.extend(batch)
        total_pages = data.get("pages", 1)
        if page >= total_pages:
            break
        page += 1
    return products


def update_product_image(api_key, api_secret, product, image_url):
    product_copy = dict(product)
    product_copy["thumbnailUrl"] = image_url
    product_copy["thumbnailURL"] = image_url
    pid = product_copy["productId"]
    return ss_request("PUT", f"/products/{pid}", api_key, api_secret, json_data=product_copy)


# ─── Compass Closeout Enrichment ──────────────────────────────────────────────

def normalize_header(h):
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def parse_closeout_skus(file_bytes):
    """Read a Compass enrichment .xlsx and return the set of SKUs (Item Number)
    flagged Store Closeout? = Y. Same file format the RPH Review tool consumes."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError("File has no header row")

    norm_header = [normalize_header(h) for h in header]
    try:
        sku_idx = next(i for i, h in enumerate(norm_header) if h in ("item number", "item #", "item"))
    except StopIteration:
        raise ValueError("Could not find an 'Item Number' column")
    try:
        closeout_idx = next(i for i, h in enumerate(norm_header) if h.startswith("store closeout"))
    except StopIteration:
        raise ValueError("Could not find a 'Store Closeout?' column")

    closeout_skus = set()
    for row in rows:
        if not row:
            continue
        sku = row[sku_idx] if sku_idx < len(row) else None
        flag = row[closeout_idx] if closeout_idx < len(row) else None
        if sku is None:
            continue
        sku = str(sku).strip()
        if sku and str(flag or "").strip().upper() == "Y":
            closeout_skus.add(sku)
    return closeout_skus


# ─── Flask App ────────────────────────────────────────────────────────────────

app = Flask(__name__)

state = {
    "api_key": "",
    "api_secret": "",
    "all_products": [],
    "missing": [],
    "log": [],
    "closeout_skus": set(),
}


HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>ShipStation Image Tool</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #101014;
    --surface: #18181c;
    --surface-2: #222228;
    --border: #2c2c34;
    --text: #ececf0;
    --text-dim: #77778a;
    --blue: #5b8def;
    --blue-dim: rgba(91,141,239,0.12);
    --green: #34c77b;
    --green-dim: rgba(52,199,123,0.12);
    --amber: #e8a735;
    --red: #e8534a;
    --radius: 12px;
    --font: 'DM Sans', system-ui, -apple-system, sans-serif;
    --mono: 'DM Mono', 'SF Mono', monospace;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: var(--font); background: var(--bg); color: var(--text); min-height: 100vh; }
  .wrap { max-width: 720px; margin: 0 auto; padding: 32px 24px; }
  .screen { display: none; }
  .screen.active { display: block; }

  .login-card, .fetch-card, .done-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 48px 40px;
    margin-top: 80px;
  }
  .fetch-card, .done-card { text-align: center; }
  .logo {
    font-size: 13px; font-weight: 600; letter-spacing: 2px;
    text-transform: uppercase; color: var(--text-dim); margin-bottom: 8px;
  }
  .login-card h1 { font-size: 26px; font-weight: 700; letter-spacing: -0.5px; margin-bottom: 32px; }
  .login-card h1 span { color: var(--blue); }
  .field { margin-bottom: 20px; }
  .field label {
    display: block; font-size: 11px; font-weight: 600; letter-spacing: 1px;
    text-transform: uppercase; color: var(--text-dim); margin-bottom: 6px;
  }
  .field input {
    width: 100%; padding: 12px 16px; background: var(--bg); border: 1px solid var(--border);
    border-radius: 8px; color: var(--text); font-family: var(--mono); font-size: 14px;
    outline: none; transition: border-color 0.2s;
  }
  .field input:focus { border-color: var(--blue); }
  .field input[type="file"] {
    padding: 9px 12px; font-family: var(--font); font-size: 13px; color: var(--text-dim);
  }
  .btn {
    display: inline-flex; align-items: center; justify-content: center; gap: 8px;
    padding: 12px 24px; font-family: var(--font); font-size: 14px; font-weight: 600;
    border: none; border-radius: 8px; cursor: pointer; transition: all 0.15s;
  }
  .btn-blue { background: var(--blue); color: white; }
  .btn-blue:hover { filter: brightness(1.1); }
  .btn-blue:disabled { opacity: 0.4; cursor: wait; }
  .btn-green { background: var(--green); color: white; }
  .btn-green:hover { filter: brightness(1.1); }
  .btn-green:disabled { opacity: 0.4; cursor: wait; }
  .btn-amber { background: var(--amber); color: #1a1a1a; }
  .btn-amber:hover { filter: brightness(1.1); }
  .btn-ghost { background: transparent; border: 1px solid var(--border); color: var(--text-dim); }
  .btn-ghost:hover { border-color: var(--blue); color: var(--blue); }
  .btn-block { width: 100%; }
  .error-msg { color: var(--red); font-size: 13px; margin-top: 12px; text-align: center; min-height: 20px; }

  .fetch-card h2, .done-card h2 { font-size: 20px; margin-bottom: 16px; }
  .fetch-card p { color: var(--text-dim); font-size: 14px; line-height: 1.6; }
  .progress-track { width: 100%; height: 4px; background: var(--border); border-radius: 2px; margin: 24px 0; overflow: hidden; }
  .progress-bar {
    height: 100%; background: var(--blue); border-radius: 2px; width: 0%;
    transition: width 0.4s;
  }
  .progress-bar.indeterminate { width: 30%; animation: ind 1.4s ease infinite; }
  @keyframes ind { 0%{transform:translateX(-100%)} 100%{transform:translateX(400%)} }

  .stats-row { display: flex; gap: 16px; margin: 24px 0; justify-content: center; }
  .stat-pill {
    background: var(--surface-2); border: 1px solid var(--border); border-radius: 8px;
    padding: 12px 20px; text-align: center; min-width: 100px;
  }
  .stat-pill .val { font-size: 28px; font-weight: 700; font-family: var(--mono); }
  .stat-pill .lbl { font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: var(--text-dim); margin-top: 2px; }

  .workflow-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 24px; padding-bottom: 16px; border-bottom: 1px solid var(--border);
  }
  .workflow-header h2 { font-size: 18px; font-weight: 600; }
  .counter {
    font-family: var(--mono); font-size: 13px; color: var(--text-dim);
    background: var(--surface-2); padding: 6px 14px; border-radius: 6px; border: 1px solid var(--border);
  }
  .counter em { font-style: normal; color: var(--blue); font-weight: 600; }

  .product-card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: var(--radius); overflow: hidden;
  }
  .product-info { padding: 16px 24px 12px; }
  .product-sku { font-family: var(--mono); font-size: 13px; font-weight: 500; color: var(--blue); margin-bottom: 4px; }
  .product-name { font-size: 15px; font-weight: 600; line-height: 1.3; margin-bottom: 10px; }
  .product-links { display: flex; gap: 10px; flex-wrap: wrap; }
  .site-link {
    font-size: 12px; font-weight: 500; color: var(--blue); text-decoration: none;
    padding: 5px 12px; border: 1px solid var(--border); border-radius: 6px; transition: all 0.15s;
  }
  .site-link:hover { border-color: var(--blue); background: var(--blue-dim); }

  .picker-wrap { padding: 0 24px 12px; }
  .picker-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 8px;
  }
  .picker-title { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; color: var(--text-dim); }
  .picker-hint { font-size: 11px; color: var(--text-dim); }
  .picker-grid {
    display: flex; gap: 8px; overflow-x: auto; padding-bottom: 6px;
  }
  .picker-grid::-webkit-scrollbar { height: 4px; }
  .picker-grid::-webkit-scrollbar-track { background: var(--border); border-radius: 2px; }
  .picker-grid::-webkit-scrollbar-thumb { background: var(--text-dim); border-radius: 2px; }
  .picker-item {
    position: relative; width: 100px; min-width: 100px; height: 100px;
    border: 2px solid var(--border);
    border-radius: 8px; overflow: hidden; cursor: pointer;
    transition: all 0.15s; background: var(--bg);
  }
  .picker-item:hover { border-color: var(--blue); transform: scale(1.03); }
  .picker-item.selected { border-color: var(--green); box-shadow: 0 0 0 2px var(--green); }
  .picker-item img { width: 100%; height: 100%; object-fit: contain; padding: 4px; }
  .picker-item.picker-match { border-color: var(--green); box-shadow: 0 0 0 1px var(--green); }
  .picker-label {
    position: absolute; bottom: 0; left: 0; right: 0;
    background: rgba(0,0,0,0.75); padding: 3px 6px;
    display: flex; align-items: center; gap: 4px; justify-content: center;
    flex-wrap: wrap;
  }
  .picker-sku { font-family: var(--mono); font-size: 10px; color: var(--text-dim); }
  .picker-badge {
    font-size: 9px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;
    padding: 1px 5px; border-radius: 3px;
  }
  .picker-badge.match { background: var(--green); color: #000; }
  .picker-loading {
    text-align: center; padding: 20px; color: var(--text-dim); font-size: 13px;
  }
  .picker-empty {
    text-align: center; padding: 16px; color: var(--text-dim); font-size: 13px;
    border: 1px dashed var(--border); border-radius: 8px;
  }

  .drop-zone-wrap { padding: 0 24px 16px; }
  .drop-zone {
    position: relative; height: 180px; border: 2px dashed var(--border); border-radius: 10px;
    display: flex; flex-direction: column; align-items: center; justify-content: center;
    gap: 8px; transition: all 0.25s; overflow: hidden; background: var(--bg);
  }
  .drop-zone.drag-over { border-color: var(--blue); background: var(--blue-dim); transform: scale(1.005); }
  .drop-zone.has-image { border-style: solid; border-color: var(--green); }
  .dz-icon {
    width: 36px; height: 36px; border-radius: 50%; background: var(--surface-2);
    display: flex; align-items: center; justify-content: center; font-size: 16px;
  }
  .dz-text { font-size: 12px; color: var(--text-dim); text-align: center; line-height: 1.4; }
  .dz-text strong { color: var(--blue); font-weight: 600; }
  .dz-hint { font-size: 11px; color: var(--text-dim); opacity: 0.6; }
  .drop-zone img {
    position: absolute; inset: 0; width: 100%; height: 100%;
    object-fit: contain; padding: 12px;
  }
  .drop-zone .img-clear {
    position: absolute; top: 10px; right: 10px; width: 28px; height: 28px;
    border-radius: 50%; background: rgba(0,0,0,0.75); color: white; font-size: 16px;
    border: none; cursor: pointer; display: flex; align-items: center; justify-content: center;
    z-index: 2; opacity: 0; transition: opacity 0.15s;
  }
  .drop-zone:hover .img-clear { opacity: 1; }

  .paste-row { display: flex; gap: 8px; padding: 0 24px; margin-bottom: 12px; }
  .paste-row input {
    flex: 1; padding: 8px 12px; background: var(--bg); border: 1px solid var(--border);
    border-radius: 8px; color: var(--text); font-family: var(--mono); font-size: 11px; outline: none;
  }
  .paste-row input:focus { border-color: var(--blue); }

  .action-row { display: flex; gap: 10px; padding: 14px 24px; border-top: 1px solid var(--border); }
  .action-row .btn { flex: 1; }

  .done-card .check {
    width: 64px; height: 64px; border-radius: 50%; background: var(--green-dim);
    display: flex; align-items: center; justify-content: center; font-size: 32px;
    margin: 0 auto 20px;
  }
  .done-card p { color: var(--text-dim); margin-bottom: 24px; }

  .toast-wrap { position: fixed; bottom: 24px; right: 24px; z-index: 999; display: flex; flex-direction: column; gap: 8px; }
  .toast {
    padding: 12px 20px; border-radius: 8px; font-size: 13px; font-weight: 500;
    color: white; animation: slideUp 0.25s ease; max-width: 380px;
  }
  @keyframes slideUp { from{transform:translateY(20px);opacity:0} to{transform:translateY(0);opacity:1} }
  .toast-ok { background: var(--green); }
  .toast-err { background: var(--red); }
  .toast-info { background: var(--blue); }
  .spinner {
    width: 18px; height: 18px; border: 2px solid rgba(255,255,255,0.25);
    border-top-color: white; border-radius: 50%; animation: spin 0.5s linear infinite; display: inline-block;
  }
  @keyframes spin { to{transform:rotate(360deg)} }
</style>
</head>
<body>
<div class="wrap">

  <!-- SCREEN 1: LOGIN -->
  <div class="screen active" id="screen-login">
    <div class="login-card">
      <div class="logo">ShipStation</div>
      <h1>Product <span>Image Tool</span></h1>
      <div class="field">
        <label>API Key</label>
        <input type="text" id="inp-key" placeholder="Enter your ShipStation API key" autocomplete="off">
      </div>
      <div class="field">
        <label>API Secret</label>
        <input type="password" id="inp-secret" placeholder="Enter your ShipStation API secret" autocomplete="off">
      </div>
      <div class="field">
        <label>Compass Closeout Export (optional)</label>
        <input type="file" id="inp-enrichment" accept=".xlsx">
        <div style="font-size:11px;color:var(--text-dim);margin-top:6px;line-height:1.5;">
          Same weekly Compass export RPH Review uses. Any SKU flagged <strong>Store Closeout? = Y</strong>
          is auto-filtered out before it reaches your queue.
        </div>
      </div>
      <button class="btn btn-blue btn-block" id="btn-connect" onclick="doConnect()">Connect</button>
      <div class="error-msg" id="login-err"></div>
    </div>
  </div>

  <!-- SCREEN 2: FETCHING -->
  <div class="screen" id="screen-fetch">
    <div class="fetch-card">
      <h2>Scanning ShipStation</h2>
      <p id="fetch-status">Fetching all products and checking for missing images...</p>
      <div class="progress-track"><div class="progress-bar indeterminate" id="fetch-progress"></div></div>
    </div>
  </div>

  <!-- SCREEN 3: RESULTS -->
  <div class="screen" id="screen-results">
    <div class="fetch-card">
      <h2>Scan Complete</h2>
      <div class="stats-row">
        <div class="stat-pill"><div class="val" id="stat-total" style="color:var(--text)">0</div><div class="lbl">Total</div></div>
        <div class="stat-pill"><div class="val" id="stat-have" style="color:var(--green)">0</div><div class="lbl">Have Images</div></div>
        <div class="stat-pill"><div class="val" id="stat-remaining" style="color:var(--amber)">0</div><div class="lbl">Remaining</div></div>
      </div>
      <p id="resume-note" style="color:var(--text-dim);font-size:13px;margin-bottom:20px;"></p>
      <div style="display:flex; gap:10px; justify-content:center; flex-wrap:wrap;">
        <button class="btn btn-blue" onclick="startAutoPass()">Start Auto-Match</button>
        <button class="btn btn-ghost" onclick="downloadXlsx()">Download XLSX</button>
        <button class="btn btn-ghost" onclick="resetProgress()" style="color:var(--red);border-color:var(--red);">Reset Progress</button>
      </div>
    </div>
  </div>

  <!-- SCREEN 3b: AUTO-PASS -->
  <div class="screen" id="screen-auto">
    <div class="fetch-card">
      <h2>Auto-Matching Images</h2>
      <p id="auto-status">Searching your website for matching product images...</p>
      <div class="progress-track"><div class="progress-bar" id="auto-progress" style="width:0%"></div></div>
      <div class="stats-row">
        <div class="stat-pill"><div class="val" id="auto-pushed" style="color:var(--green)">0</div><div class="lbl">Auto-Pushed</div></div>
        <div class="stat-pill"><div class="val" id="auto-closeout" style="color:var(--amber)">0</div><div class="lbl">Auto-Closeout</div></div>
        <div class="stat-pill"><div class="val" id="auto-manual" style="color:var(--blue)">0</div><div class="lbl">Need Manual</div></div>
      </div>
      <p id="auto-current" style="color:var(--text-dim);font-size:12px;font-family:var(--mono);margin-top:8px;"></p>
      <button class="btn btn-blue" id="btn-review-now" onclick="startManualPhase()" style="display:none;margin-top:20px;">Review items now</button>
    </div>
  </div>

  <!-- SCREEN 4: ONE-BY-ONE -->
  <div class="screen" id="screen-work">
    <div class="workflow-header">
      <h2>Assign Image</h2>
      <div style="display:flex;gap:8px;align-items:center;">
        <span class="picker-hint" id="work-live-note" style="display:none;">scanning for more&hellip;</span>
        <button class="btn btn-ghost" onclick="downloadXlsx()" style="padding:8px 14px;font-size:12px;">XLSX</button>
        <div class="counter"><em id="work-cur">1</em> / <span id="work-total">0</span></div>
      </div>
    </div>
    <div class="product-card" id="waiting-card" style="display:none;">
      <div class="fetch-card" style="margin-top:0;border:none;padding:48px 40px;">
        <span class="spinner" style="width:28px;height:28px;border-color:rgba(91,141,239,0.25);border-top-color:var(--blue);"></span>
        <p style="margin-top:16px;">You're caught up. Auto-match is still scanning in the background &mdash; more items will appear here as they're found.</p>
      </div>
    </div>
    <div class="product-card" id="work-card">
      <div class="product-info">
        <div class="product-sku" id="work-sku"></div>
        <div class="product-name" id="work-name"></div>
        <div class="product-links" id="work-links"></div>
      </div>

      <!-- Image picker from site scrape -->
      <div class="picker-wrap" id="picker-wrap" style="display:none;">
        <div class="picker-header">
          <span class="picker-title">Images found on site</span>
          <span class="picker-hint" id="picker-hint"></span>
        </div>
        <div class="picker-grid" id="picker-grid"></div>
      </div>

      <div class="drop-zone-wrap">
        <div class="drop-zone" id="drop-zone"
             ondragenter="dzEnter(event)" ondragover="dzOver(event)"
             ondragleave="dzLeave(event)" ondrop="dzDrop(event)">
          <div class="dz-placeholder">
            <div class="dz-icon">🖼</div>
            <div class="dz-text">Select an image above, drag one here, or paste a URL below</div>
            <div class="dz-hint">or paste the image URL below</div>
          </div>
        </div>
      </div>
      <div class="paste-row">
        <input type="text" id="paste-url" placeholder="Paste image URL here..." onkeydown="if(event.key==='Enter')applyPaste()">
        <button class="btn btn-ghost" onclick="applyPaste()" style="flex:none;">Set</button>
      </div>
      <div class="action-row">
        <button class="btn btn-ghost" id="btn-skip" onclick="doSkip()">Skip</button>
        <button class="btn btn-amber" id="btn-closeout" onclick="doCloseout()">Closeout</button>
        <button class="btn btn-green" id="btn-push" onclick="doPush()" disabled>Push to ShipStation</button>
      </div>
    </div>
  </div>

  <!-- SCREEN 5: DONE -->
  <div class="screen" id="screen-done">
    <div class="done-card">
      <div class="check">✓</div>
      <h2>All Done</h2>
      <p id="done-summary"></p>
      <div style="display:flex;gap:10px;justify-content:center;flex-wrap:wrap;">
        <button class="btn btn-ghost" onclick="downloadXlsx()">Download XLSX</button>
        <button class="btn btn-ghost" onclick="location.reload()">Start Over</button>
      </div>
    </div>
  </div>
</div>

<div class="toast-wrap" id="toasts"></div>

<script>
let currentUrl = '';
let pushed = 0;
let skipped = 0;
let closeouts = 0;

function openSite(url) {
  window.open(url, 'cheshire_site', 'noopener');
}

function show(id) {
  document.querySelectorAll('.screen').forEach(s => s.classList.remove('active'));
  document.getElementById(id).classList.add('active');
}
function toast(msg, type) {
  const t = document.createElement('div');
  t.className = 'toast toast-' + (type||'info');
  t.textContent = msg;
  document.getElementById('toasts').appendChild(t);
  setTimeout(() => t.remove(), 3500);
}
function esc(s) { const d = document.createElement('div'); d.textContent = s||''; return d.innerHTML; }

async function doConnect() {
  const key = document.getElementById('inp-key').value.trim();
  const secret = document.getElementById('inp-secret').value.trim();
  const enrichFile = document.getElementById('inp-enrichment').files[0];
  const err = document.getElementById('login-err');
  const btn = document.getElementById('btn-connect');
  if (!key || !secret) { err.textContent = 'Enter both API key and secret.'; return; }
  btn.disabled = true; btn.innerHTML = '<span class="spinner"></span>'; err.textContent = '';
  try {
    const r = await fetch('/api/connect', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({api_key:key, api_secret:secret})});
    const d = await r.json();
    if (!r.ok) throw new Error(d.error||'Failed');

    if (enrichFile) {
      const fd = new FormData();
      fd.append('file', enrichFile);
      const ur = await fetch('/api/upload_enrichment', {method:'POST', body: fd});
      const ud = await ur.json();
      if (!ur.ok) throw new Error('Closeout export: ' + (ud.error || 'upload failed'));
    }

    show('screen-fetch'); doFetch();
  } catch(e) { err.textContent = e.message; btn.disabled = false; btn.textContent = 'Connect'; }
}

async function doFetch() {
  try {
    const r = await fetch('/api/fetch');
    const d = await r.json();
    if (!r.ok) throw new Error(d.error);
    document.getElementById('stat-total').textContent = d.total;
    document.getElementById('stat-have').textContent = d.have_images;
    document.getElementById('stat-remaining').textContent = d.remaining;
    const note = document.getElementById('resume-note');
    const parts = [];
    if (d.already_handled > 0) {
      parts.push('Resumed from saved progress — ' + d.already_handled + ' products already handled (' + d.missing_total + ' were missing total).');
    } else {
      parts.push(d.missing_total + ' products missing images.');
    }
    if (d.auto_closeout_enrichment > 0) {
      parts.push(d.auto_closeout_enrichment + ' auto-filtered as closeouts from your Compass export.');
    }
    note.textContent = parts.join(' ');
    show('screen-results');
  } catch(e) {
    document.getElementById('fetch-status').textContent = 'Error: ' + e.message;
    document.getElementById('fetch-progress').classList.remove('indeterminate');
  }
}

// ── Phase 1: Auto-pass (processes products concurrently in the background) ──
const AUTO_MATCH_CONCURRENCY = 6;
let manualIndices = [];
let manualPos = 0;
let autoPassRunning = false;
let manualReviewActive = false;
let waitingPollTimer = null;

function updateManualCounterUI() {
  document.getElementById('auto-manual').textContent = manualIndices.length;
  if (manualReviewActive) return;
  const btn = document.getElementById('btn-review-now');
  if (manualIndices.length > 0) {
    btn.style.display = 'inline-flex';
    btn.textContent = 'Review ' + manualIndices.length + ' item' + (manualIndices.length === 1 ? '' : 's') + ' now';
  }
}

async function startAutoPass() {
  pushed = 0; skipped = 0; closeouts = 0;
  manualIndices = [];
  autoPassRunning = true;
  manualReviewActive = false;
  document.getElementById('btn-review-now').style.display = 'none';
  show('screen-auto');

  const resp = await fetch('/api/product_count');
  const countData = await resp.json();
  const total = countData.count;

  if (total === 0) { autoPassRunning = false; finishUp(); return; }

  let nextIndex = 0;
  let completed = 0;

  async function processOne(i) {
    const pResp = await fetch('/api/product/' + i);
    const p = await pResp.json();
    if (p.done) return;

    if (!p.sku) {
      manualIndices.push(i);
      updateManualCounterUI();
      return;
    }

    try {
      const sResp = await fetch('/api/scrape?q=' + encodeURIComponent(p.sku) + '&sku=' + encodeURIComponent(p.sku));
      const s = await sResp.json();

      if (s.images && s.images.length > 0 && s.images[0].exact_match) {
        // Auto-push
        const pushResp = await fetch('/api/push', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index: i, image_url: s.images[0].url})});
        if (pushResp.ok) {
          pushed++;
          document.getElementById('auto-pushed').textContent = pushed;
        } else {
          manualIndices.push(i);
          updateManualCounterUI();
        }
      } else if (!s.images || s.images.length === 0) {
        // No images — auto-closeout
        await fetch('/api/closeout_at', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index: i})});
        closeouts++;
        document.getElementById('auto-closeout').textContent = closeouts;
      } else {
        // Images but no exact match — needs manual
        manualIndices.push(i);
        updateManualCounterUI();
      }
    } catch(e) {
      manualIndices.push(i);
      updateManualCounterUI();
    }
  }

  async function worker() {
    while (nextIndex < total) {
      const i = nextIndex++;
      await processOne(i);
      completed++;
      const pct = Math.round((completed / total) * 100);
      document.getElementById('auto-progress').style.width = pct + '%';
      document.getElementById('auto-current').textContent = completed + ' / ' + total + ' scanned';
    }
  }

  const workers = [];
  for (let w = 0; w < Math.min(AUTO_MATCH_CONCURRENCY, total); w++) workers.push(worker());
  await Promise.all(workers);

  // Auto-pass done
  autoPassRunning = false;
  document.getElementById('auto-current').textContent = '';

  if (manualReviewActive) {
    // User is already reviewing — the manual screen's own polling will notice
    // autoPassRunning is now false and finish up once they're caught up.
    toast('Background auto-match finished.', 'info');
    return;
  }

  if (manualIndices.length === 0) {
    document.getElementById('auto-status').textContent = 'All done — no manual work needed!';
    setTimeout(() => finishUp(), 1500);
  } else {
    document.getElementById('auto-status').textContent = 'Auto-match complete! ' + manualIndices.length + ' products need your help.';
    setTimeout(() => startManualPhase(), 1500);
  }
}

// ── Phase 2: Manual workflow. Can start while auto-match is still running in
// the background (via "Review Now"), so manualIndices may keep growing while
// the user works through it — loadManualProduct() handles the catch-up case.
function startManualPhase() {
  manualReviewActive = true;
  manualPos = 0;
  show('screen-work');
  loadManualProduct();
}

function loadManualProduct() {
  document.getElementById('work-live-note').style.display = autoPassRunning ? 'inline' : 'none';

  if (manualPos >= manualIndices.length) {
    if (autoPassRunning) { showWaitingForMore(); return; }
    finishUp();
    return;
  }

  clearTimeout(waitingPollTimer);
  document.getElementById('waiting-card').style.display = 'none';
  document.getElementById('work-card').style.display = 'block';

  clearDZ();
  document.getElementById('picker-wrap').style.display = 'none';
  document.getElementById('picker-grid').innerHTML = '';

  const idx = manualIndices[manualPos];
  fetch('/api/product/' + idx).then(r=>r.json()).then(d => {
    if (d.done) { manualPos++; loadManualProduct(); return; }
    document.getElementById('work-cur').textContent = (manualPos + 1);
    document.getElementById('work-total').textContent = manualIndices.length;
    document.getElementById('work-sku').textContent = d.sku || 'No SKU';
    document.getElementById('work-name').textContent = d.name || 'Unnamed Product';
    const sku = encodeURIComponent(d.sku||'');
    document.getElementById('work-links').innerHTML =
      '<a class="site-link" href="#" onclick="openSite(\'https://www.cheshirehorse.com/search?q='+sku+'\');return false;">🔍 Search by SKU</a>';
    document.getElementById('btn-push').disabled = true;
    document.getElementById('btn-push').textContent = 'Push to ShipStation';

    if (d.sku) {
      scrapeImagesManual(d.sku);
    }
  });
}

function showWaitingForMore() {
  document.getElementById('work-card').style.display = 'none';
  document.getElementById('waiting-card').style.display = 'block';
  document.getElementById('work-cur').textContent = manualIndices.length;
  document.getElementById('work-total').textContent = manualIndices.length;
  clearTimeout(waitingPollTimer);
  waitingPollTimer = setTimeout(() => { if (manualReviewActive) loadManualProduct(); }, 900);
}

function scrapeImagesManual(sku) {
  const wrap = document.getElementById('picker-wrap');
  const grid = document.getElementById('picker-grid');
  const hint = document.getElementById('picker-hint');
  wrap.style.display = 'block';
  grid.innerHTML = '<div class="picker-loading"><span class="spinner"></span> Searching site &amp; product pages...</div>';
  hint.textContent = '';

  fetch('/api/scrape?q=' + encodeURIComponent(sku) + '&sku=' + encodeURIComponent(sku))
    .then(r => r.json())
    .then(d => {
      if (d.images && d.images.length > 0) {
        renderPicker(d.images, sku);
        hint.textContent = d.images.length + ' image(s) — pick the right one';
      } else {
        grid.innerHTML = '<div class="picker-empty">No images found. Drag from the website or paste a URL.</div>';
      }
    })
    .catch(e => {
      grid.innerHTML = '<div class="picker-empty">Could not search site: ' + esc(e.message) + '</div>';
    });
}

function renderPicker(images, targetSku) {
  const grid = document.getElementById('picker-grid');
  const skuLower = (targetSku||'').toLowerCase();
  grid.innerHTML = images.map((img, i) => {
    const fileSku = img.file_sku || '';
    const isMatch = fileSku && skuLower && fileSku.toLowerCase() === skuLower;
    const matchClass = isMatch ? ' picker-match' : '';
    const label = fileSku ? fileSku : '';
    const badge = isMatch ? '<span class="picker-badge match">SKU Match</span>' :
                  (img.exact_match ? '<span class="picker-badge match">Match</span>' : '');
    return '<div class="picker-item' + matchClass + '" onclick="pickImage(this, \'' + img.url.replace(/'/g, "\\'") + '\')" title="' + esc(img.alt || fileSku) + '">' +
      '<img src="' + esc(img.url) + '" alt="' + esc(img.alt) + '" onerror="this.parentElement.style.display=\'none\'">' +
      (label || badge ? '<div class="picker-label">' + badge + (label ? '<span class="picker-sku">' + esc(label) + '</span>' : '') + '</div>' : '') +
      '</div>';
  }).join('');
}

function pickImage(el, url) {
  // Deselect all, select this one
  document.querySelectorAll('.picker-item').forEach(p => p.classList.remove('selected'));
  el.classList.add('selected');
  setImage(url);
}

function dzEnter(e) { e.preventDefault(); e.currentTarget.classList.add('drag-over'); }
function dzOver(e) { e.preventDefault(); }
function dzLeave(e) { e.currentTarget.classList.remove('drag-over'); }
function dzDrop(e) {
  e.preventDefault();
  document.getElementById('drop-zone').classList.remove('drag-over');
  let url = e.dataTransfer.getData('text/uri-list')||'';
  if (!url) url = e.dataTransfer.getData('text/plain')||'';
  if (!url) { const h = e.dataTransfer.getData('text/html')||''; const m = h.match(/src=["']([^"']+)["']/i); if(m) url=m[1]; }
  url = url.trim();
  if (url && /^https?:\/\//.test(url)) {
    setImage(url);
    currentUrl = url;
    doPush();
  }
  else { toast('Could not get image URL — try pasting instead.','err'); }
}

function setImage(url) {
  currentUrl = url;
  const dz = document.getElementById('drop-zone');
  dz.classList.add('has-image');
  dz.innerHTML = '<img src="'+esc(url)+'" alt="Product" onerror="this.style.display=\'none\'">' +
    '<button class="img-clear" onclick="clearDZ()" title="Remove">&times;</button>';
  document.getElementById('paste-url').value = url;
  document.getElementById('btn-push').disabled = false;
}
function clearDZ() {
  currentUrl = '';
  const dz = document.getElementById('drop-zone');
  dz.classList.remove('has-image');
  dz.innerHTML = '<div class="dz-placeholder"><div class="dz-icon">🖼</div>' +
    '<div class="dz-text">Select an image above, drag one here, or paste a URL below</div>' +
    '<div class="dz-hint">or paste the image URL below</div></div>';
  document.getElementById('paste-url').value = '';
  document.getElementById('btn-push').disabled = true;
}
function applyPaste() {
  const url = document.getElementById('paste-url').value.trim();
  if (url && /^https?:\/\//.test(url)) setImage(url);
  else toast('Enter a valid URL starting with http:// or https://','err');
}

async function doPush() {
  const btn = document.getElementById('btn-push');
  btn.disabled = true; btn.innerHTML = '<span class="spinner"></span> Pushing...';
  try {
    const idx = manualIndices[manualPos];
    const r = await fetch('/api/push', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index: idx, image_url:currentUrl})});
    const d = await r.json();
    if (!r.ok) throw new Error(d.error);
    pushed++;
    toast('Image pushed!','ok');
    manualPos++;
    loadManualProduct();
  } catch(e) { toast('Push failed: '+e.message,'err'); btn.disabled=false; btn.textContent='Push to ShipStation'; }
}
function doSkip() {
  const idx = manualIndices[manualPos];
  fetch('/api/skip',{method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index: idx})}).then(r=>r.json()).then(d => {
    skipped++;
    manualPos++;
    loadManualProduct();
  });
}
function doCloseout() {
  const idx = manualIndices[manualPos];
  fetch('/api/closeout',{method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({index: idx})}).then(r=>r.json()).then(d => {
    closeouts++;
    toast('Marked as closeout','info');
    manualPos++;
    loadManualProduct();
  });
}
function finishUp() {
  let parts = [];
  if (pushed) parts.push(pushed + ' pushed');
  if (closeouts) parts.push(closeouts + ' closeouts');
  if (skipped) parts.push(skipped + ' skipped');
  document.getElementById('done-summary').textContent = parts.join(', ') + '.';
  show('screen-done');
}
function downloadXlsx() { window.open('/api/xlsx','_blank'); }
function resetProgress() {
  if (!confirm('This will clear all saved progress (skips, closeouts, pushes). Are you sure?')) return;
  fetch('/api/reset',{method:'POST'}).then(r=>r.json()).then(d => {
    toast('Progress reset — re-scanning...','info');
    show('screen-fetch'); doFetch();
  });
}
</script>
</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/connect", methods=["POST"])
def api_connect():
    data = request.json
    state["api_key"] = data.get("api_key", "")
    state["api_secret"] = data.get("api_secret", "")
    try:
        ss_request("GET", "/products", state["api_key"], state["api_secret"],
                    params={"page": 1, "pageSize": 1})
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 401:
            return jsonify({"error": "Invalid credentials — check your API key and secret."}), 401
        return jsonify({"error": f"Connection error: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Connection error: {e}"}), 500
    return jsonify({"ok": True})


@app.route("/api/fetch")
def api_fetch():
    try:
        all_prods = fetch_all_products(state["api_key"], state["api_secret"])
        state["all_products"] = all_prods
        missing = [p for p in all_prods if not p.get("thumbnailUrl")]

        # Load saved progress
        saved = load_progress()
        state["log"] = saved.get("log", [])

        # Build set of SKUs already handled (pushed, skipped, closeout)
        handled_skus = {e["sku"] for e in state["log"] if e.get("sku")}

        # Auto-closeout any missing product flagged Store Closeout? = Y in the
        # uploaded Compass enrichment export, before it ever reaches the queue.
        closeout_skus = state.get("closeout_skus") or set()
        auto_closeout_count = 0
        if closeout_skus:
            for p in missing:
                sku = p.get("sku", "")
                if sku and sku in closeout_skus and sku not in handled_skus:
                    state["log"].append({
                        "sku": sku,
                        "name": p.get("name", ""),
                        "url": "",
                        "action": "closeout",
                        "source": "enrichment",
                        "timestamp": datetime.now().isoformat(),
                    })
                    handled_skus.add(sku)
                    auto_closeout_count += 1
            if auto_closeout_count:
                save_progress()

        # Filter missing to only unhandled products
        remaining = [p for p in missing if p.get("sku", "") not in handled_skus]

        state["missing"] = remaining

        return jsonify({
            "total": len(all_prods),
            "have_images": len(all_prods) - len(missing),
            "missing_total": len(missing),
            "already_handled": len(handled_skus),
            "auto_closeout_enrichment": auto_closeout_count,
            "remaining": len(remaining),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload_enrichment", methods=["POST"])
def api_upload_enrichment():
    """Accept a Compass enrichment .xlsx and store its closeout SKUs for
    filtering on the next /api/fetch."""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        closeout_skus = parse_closeout_skus(file.read())
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    state["closeout_skus"] = closeout_skus
    return jsonify({"ok": True, "closeout_count": len(closeout_skus)})


@app.route("/api/product/<int:idx>")
def api_product(idx):
    if idx >= len(state["missing"]):
        return jsonify({"done": True})
    p = state["missing"][idx]
    return jsonify({
        "sku": p.get("sku", ""),
        "name": p.get("name", ""),
        "product_id": p["productId"],
        "total": len(state["missing"]),
        "done": False,
    })


@app.route("/api/product_count")
def api_product_count():
    return jsonify({"count": len(state["missing"])})


@app.route("/api/closeout_at", methods=["POST"])
def api_closeout_at():
    """Closeout a product at a specific index (used during auto-pass)."""
    data = request.json
    idx = data.get("index", 0)
    if idx is None or idx >= len(state["missing"]):
        return jsonify({"error": "Index out of range"}), 400
    product = state["missing"][idx]
    log_action(product, "closeout")
    return jsonify({"ok": True})


@app.route("/api/push", methods=["POST"])
def api_push():
    # index is passed explicitly (rather than relying on shared server state)
    # because auto-match (background) and manual review can push concurrently.
    data = request.json
    image_url = data["image_url"]
    idx = data.get("index")
    if idx is None or idx >= len(state["missing"]):
        return jsonify({"error": "Invalid index"}), 400
    product = state["missing"][idx]
    try:
        update_product_image(state["api_key"], state["api_secret"], product, image_url)
        log_action(product, "pushed", image_url=image_url)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/skip", methods=["POST"])
def api_skip():
    data = request.json or {}
    idx = data.get("index")
    if idx is None or idx >= len(state["missing"]):
        return jsonify({"error": "Invalid index"}), 400
    product = state["missing"][idx]
    log_action(product, "skipped")
    return jsonify({"ok": True})


@app.route("/api/closeout", methods=["POST"])
def api_closeout():
    data = request.json or {}
    idx = data.get("index")
    if idx is None or idx >= len(state["missing"]):
        return jsonify({"error": "Invalid index"}), 400
    product = state["missing"][idx]
    log_action(product, "closeout")
    return jsonify({"ok": True})


@app.route("/api/scrape")
def api_scrape():
    """Fetch cheshirehorse.com search results, follow to PDP, and extract all product/variant images."""
    from bs4 import BeautifulSoup

    query = request.args.get("q", "").strip()
    sku = request.args.get("sku", "").strip()  # The ShipStation variant SKU to try to match
    if not query:
        return jsonify({"images": [], "error": "No query provided"})

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    def normalize_img_url(src):
        """Make URL absolute and upscale to 650px."""
        if not src:
            return ""
        if src.startswith("//"):
            src = "https:" + src
        elif src.startswith("/"):
            src = "https://www.cheshirehorse.com" + src
        if "sw=" in src:
            src = re.sub(r"sw=\d+", "sw=650", src)
        if "sh=" in src:
            src = re.sub(r"sh=\d+", "sh=650", src)
        return src

    def is_product_image(src):
        return src and "demandware.static" in src and "images/products" in src

    def extract_sku_from_url(url):
        """Try to extract a SKU/product ID from the image filename."""
        # Pattern: .../images/products/030336.jpg or .../images/products/030336-1.jpg
        m = re.search(r'/images/products/([^/?]+?)\.\w+', url)
        if m:
            return m.group(1)
        return ""

    def is_sku_match(file_sku, target_sku):
        """Check if file_sku matches the target — exact or with a -1 suffix."""
        if file_sku == target_sku:
            return True
        if file_sku == target_sku + '-1':
            return True
        return False

    def extract_images_from_soup(soup):
        """Pull all product images from a page's HTML."""
        found = []
        seen = set()

        # img tags (src, data-src, data-lazy)
        for img in soup.find_all("img"):
            for attr in ["src", "data-src", "data-lazy"]:
                src = img.get(attr, "")
                if is_product_image(src):
                    src = normalize_img_url(src)
                    if src not in seen:
                        seen.add(src)
                        found.append({
                            "url": src,
                            "alt": img.get("alt", ""),
                            "file_sku": extract_sku_from_url(src),
                        })
                    break

        # og:image meta
        for meta in soup.find_all("meta", property="og:image"):
            src = meta.get("content", "")
            if is_product_image(src):
                src = normalize_img_url(src)
                if src not in seen:
                    seen.add(src)
                    found.append({"url": src, "alt": "og:image", "file_sku": extract_sku_from_url(src)})

        # JSON-LD product data (SFCC often embeds variant images here)
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "")
                items = data if isinstance(data, list) else [data]
                for item in items:
                    if item.get("@type") in ("Product", "IndividualProduct"):
                        for img_url in (item.get("image") if isinstance(item.get("image"), list) else [item.get("image", "")]):
                            if img_url and is_product_image(img_url):
                                img_url = normalize_img_url(img_url)
                                if img_url not in seen:
                                    seen.add(img_url)
                                    found.append({"url": img_url, "alt": item.get("name", ""), "file_sku": extract_sku_from_url(img_url)})
                        for offer in (item.get("offers", []) if isinstance(item.get("offers"), list) else [item.get("offers", {})]):
                            img_url = offer.get("image", "")
                            if img_url and is_product_image(img_url):
                                img_url = normalize_img_url(img_url)
                                if img_url not in seen:
                                    seen.add(img_url)
                                    found.append({"url": img_url, "alt": offer.get("name", ""), "file_sku": extract_sku_from_url(img_url)})
            except (json.JSONDecodeError, TypeError, AttributeError):
                pass

        # data-imgs or similar JSON attributes
        for el in soup.find_all(attrs={"data-imgs": True}):
            try:
                imgs_data = json.loads(el.get("data-imgs", "{}"))
                for key, val in imgs_data.items():
                    urls = val if isinstance(val, list) else [val]
                    for u in urls:
                        src = u.get("url", u) if isinstance(u, dict) else u
                        if is_product_image(str(src)):
                            src = normalize_img_url(str(src))
                            if src not in seen:
                                seen.add(src)
                                found.append({"url": src, "alt": "", "file_sku": extract_sku_from_url(src)})
            except (json.JSONDecodeError, TypeError):
                pass

        # Background images in style attrs
        for el in soup.find_all(style=True):
            style = el.get("style", "")
            m = re.search(r'url\(["\']?([^"\')\s]+demandware[^"\')\s]+images/products[^"\')\s]+)["\']?\)', style)
            if m:
                src = normalize_img_url(m.group(1))
                if src not in seen:
                    seen.add(src)
                    found.append({"url": src, "alt": "", "file_sku": extract_sku_from_url(src)})

        return found

    def find_product_links(soup):
        """Find links to product detail pages from search results."""
        links = []
        seen = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/p/" in href or "/product/" in href:
                if href.startswith("/"):
                    href = "https://www.cheshirehorse.com" + href
                path = href.split("?")[0]
                if path not in seen:
                    seen.add(path)
                    links.append(href)
        return links[:5]

    # Step 1: Fetch search results page
    search_url = f"https://www.cheshirehorse.com/search?q={query}"
    try:
        resp = scrape_session.get(search_url, timeout=15, headers=headers)
        resp.raise_for_status()
    except Exception as e:
        return jsonify({"images": [], "error": f"Could not fetch search page: {e}"})

    soup = BeautifulSoup(resp.text, "html.parser")

    # Get images from search results page itself
    images = extract_images_from_soup(soup)

    # Step 2: Follow through to PDPs to get variant images. Fetched concurrently
    # (rather than one at a time) since this is the slow part of each scrape —
    # up to 5 page fetches per product otherwise done in series.
    def fetch_pdp_images(pdp_url):
        try:
            pdp_resp = scrape_session.get(pdp_url, timeout=15, headers=headers)
            pdp_resp.raise_for_status()
            pdp_soup = BeautifulSoup(pdp_resp.text, "html.parser")
            return extract_images_from_soup(pdp_soup)
        except Exception:
            return []

    pdp_links = find_product_links(soup)
    if pdp_links:
        existing_urls = {img["url"] for img in images}
        with ThreadPoolExecutor(max_workers=min(3, len(pdp_links))) as executor:
            futures = [executor.submit(fetch_pdp_images, url) for url in pdp_links]
            for future in as_completed(futures):
                for img in future.result():
                    if img["url"] not in existing_urls:
                        images.append(img)
                        existing_urls.add(img["url"])

    # Step 3: Sort images — exact SKU match (including -1 suffix) first
    if sku:
        sku_lower = sku.lower().strip()
        def match_score(img):
            file_sku = (img.get("file_sku") or "").lower()
            if is_sku_match(file_sku, sku_lower):
                return 0  # Exact or suffix match — best
            if sku_lower in file_sku or file_sku in sku_lower:
                return 1  # Partial match
            return 2  # No match
        images.sort(key=match_score)

        # Mark the best match
        if images and match_score(images[0]) == 0:
            images[0]["exact_match"] = True

    return jsonify({"images": images, "query": query, "sku": sku})


@app.route("/api/reset", methods=["POST"])
def api_reset():
    state["log"] = []
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
    return jsonify({"ok": True})


@app.route("/api/xlsx")
def api_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Images"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2B3A5C")
    hdr_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Arial", size=11)
    link_font = Font(name="Arial", size=11, color="3366CC", underline="single")
    alt_fill = PatternFill("solid", fgColor="F5F7FA")
    thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))

    headers = ["SKU", "Product Name", "Product ID", "Search on Site", "Status"]
    widths = [20, 50, 15, 55, 12]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        ws.column_dimensions[chr(64 + col) if col < 27 else "E"].width = w

    # Set column widths properly for all columns
    for col, w in enumerate(widths, 1):
        col_letter = chr(64 + col)
        ws.column_dimensions[col_letter].width = w

    pushed_skus = {e["sku"] for e in state["log"] if e["action"] == "pushed"}
    closeout_skus = {e["sku"] for e in state["log"] if e["action"] == "closeout"}
    skipped_skus = {e["sku"] for e in state["log"] if e["action"] == "skipped"}

    # Include ALL products missing images (current remaining + already handled)
    all_missing = [p for p in state["all_products"] if not p.get("thumbnailUrl")]

    for i, p in enumerate(all_missing):
        row = i + 2
        sku = p.get("sku", "")
        name = p.get("name", "")
        pid = p.get("productId", "")
        search_url = f"https://www.cheshirehorse.com/search?q={sku}"

        if sku in pushed_skus:
            status = "Pushed"
        elif sku in closeout_skus:
            status = "Closeout"
        elif sku in skipped_skus:
            status = "Skipped"
        else:
            status = "Missing"

        ws.cell(row=row, column=1, value=sku).font = cell_font
        ws.cell(row=row, column=2, value=name).font = cell_font
        ws.cell(row=row, column=3, value=pid).font = cell_font

        lc = ws.cell(row=row, column=4, value=search_url)
        lc.font = link_font
        lc.hyperlink = search_url

        sc = ws.cell(row=row, column=5, value=status)
        sc.alignment = Alignment(horizontal="center")
        if status == "Pushed":
            sc.font = Font(name="Arial", size=11, color="1B7A3D", bold=True)
        elif status == "Closeout":
            sc.font = Font(name="Arial", size=11, color="C45500", bold=True)
        elif status == "Skipped":
            sc.font = Font(name="Arial", size=11, color="666666")
        else:
            sc.font = Font(name="Arial", size=11, color="B8860B")

        if i % 2 == 1:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = alt_fill
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = thin_border

    ws.auto_filter.ref = f"A1:E{len(all_missing) + 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"shipstation_missing_images_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


def main():
    parser = argparse.ArgumentParser(description="ShipStation Product Image Tool v2")
    parser.add_argument("--port", type=int, default=5050)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--no-browser", action="store_true", help="Don't auto-open a browser tab")
    args = parser.parse_args()

    url = f"http://{args.host}:{args.port}"

    print("=" * 52)
    print("  ShipStation Image Tool v2")
    print("=" * 52)
    print(f"  Open {url} in your browser")
    print(f"  Press Ctrl+C to stop")
    print("=" * 52)

    if not args.no_browser:
        Timer(1.0, lambda: webbrowser.open(url)).start()

    # threaded=True so the browser can run auto-match (concurrent scrape
    # requests) and manual review requests at the same time.
    app.run(host=args.host, port=args.port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
